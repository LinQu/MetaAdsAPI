import csv
import json
import os
import sqlite3
import tempfile
import unittest
from datetime import date
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from meta_ads_tool.api import MetaAdsClient, MetaApiError
from meta_ads_tool.checkpoint import CheckpointMismatchError, CheckpointStore
from meta_ads_tool.cli import main as cli_main
from meta_ads_tool.constants import CSV_REPORT_COLUMNS, INPUT_HEADERS
from meta_ads_tool.dates import preset_date_range, resolve_date_range
from meta_ads_tool.exporters import export_csv, export_xlsx, sanitize_spreadsheet_text
from meta_ads_tool.input_excel import read_accounts_from_excel
from meta_ads_tool.runner import process_rinci_account, run_accounts
from meta_ads_tool.transform import normalize_ad_name, parse_hour_bucket, time_category


class FakeResponse:
    def __init__(self, status_code, payload, text=""):
        self.status_code = status_code
        self._payload = payload
        self.text = text
        self.ok = 200 <= status_code < 300
        self.headers = {}

    def json(self):
        if isinstance(self._payload, Exception):
            raise self._payload
        return self._payload


class FakeSession:
    def __init__(self, responses):
        self.responses = list(responses)
        self.calls = 0
        self.requests = []
        self.headers = {}

    def get(self, *args, **kwargs):
        self.requests.append((args, kwargs))
        response = self.responses[self.calls]
        self.calls += 1
        return response


class DateTests(unittest.TestCase):
    def test_presets(self):
        current = date(2026, 8, 5)
        self.assertEqual(
            preset_date_range("7-hari-terakhir", today=current),
            (date(2026, 7, 30), date(2026, 8, 5)),
        )
        self.assertEqual(
            preset_date_range("bulan-lalu", today=current),
            (date(2026, 7, 1), date(2026, 7, 31)),
        )

    def test_resolve_manual(self):
        since, until, source = resolve_date_range(
            "2026-08-01", "2026-08-05", None
        )
        self.assertEqual(
            (since, until, source),
            ("2026-08-01", "2026-08-05", "manual"),
        )


class TransformTests(unittest.TestCase):
    def test_hour_bucket_and_category(self):
        text, hour = parse_hour_bucket("7:00:00 - 7:59:59")
        self.assertEqual(text, "07:00 - 07:59")
        self.assertEqual(hour, 7)
        self.assertEqual(time_category(hour), "Pagi")
        self.assertEqual(time_category(0), "Dini hari")
        self.assertEqual(time_category(12), "Siang")
        self.assertEqual(time_category(16), "Sore")
        self.assertEqual(time_category(22), "Malam")

    def test_formula_injection(self):
        self.assertEqual(sanitize_spreadsheet_text("=1+1"), "'=1+1")
        self.assertEqual(
            sanitize_spreadsheet_text("  @SUM(A1)"),
            "'  @SUM(A1)",
        )
        self.assertEqual(sanitize_spreadsheet_text("Nama normal"), "Nama normal")

    def test_normalize_ad_name(self):
        self.assertEqual(
            normalize_ad_name('Postingan: "Lagi butuh dana cepat tapi nggak mau"'),
            "Lagi butuh dana cepat tapi nggak mau",
        )
        self.assertEqual(
            normalize_ad_name("postingan: “Teks dengan petik tipografis”"),
            "Teks dengan petik tipografis",
        )
        self.assertEqual(normalize_ad_name("Nama iklan biasa"), "Nama iklan biasa")
        self.assertEqual(normalize_ad_name(None), "")


class InputExcelTests(unittest.TestCase):
    def test_reads_exact_four_columns_and_token(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "accounts.xlsx")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(list(INPUT_HEADERS))
            sheet.append(["123456789", "EAA-secret-token", "JKT01", "GADAI"])
            workbook.save(path)
            workbook.close()

            accounts = read_accounts_from_excel(path)
            self.assertEqual(accounts, [{
                "account_id": "123456789",
                "access_token": "EAA-secret-token",
                "cabang": "JKT01",
                "bisnis": "GADAI",
            }])

    def test_rejects_wrong_header_order(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "accounts.xlsx")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["TOKEN", "ID ACCOUNT", "CABANG", "BISNIS"])
            sheet.append(["token", "123", "CAB", "GADAI"])
            workbook.save(path)
            workbook.close()

            with self.assertRaises(RuntimeError):
                read_accounts_from_excel(path)


class RetryTests(unittest.TestCase):
    def test_retry_then_success(self):
        session = FakeSession([
            FakeResponse(
                500,
                {
                    "error": {
                        "message": "temporary",
                        "code": 2,
                        "is_transient": True,
                    }
                },
            ),
            FakeResponse(200, {"data": [{"id": "1"}]}),
        ])
        sleeps = []
        client = MetaAdsClient(
            api_version="v25.0",
            access_token="token",
            max_retries=2,
            retry_base_delay=0.1,
            retry_max_delay=1.0,
            sleep_func=sleeps.append,
            jitter_func=lambda low, high: 0.0,
            session=session,
        )
        payload = client.get_json("https://example.invalid")
        self.assertEqual(payload["data"][0]["id"], "1")
        self.assertEqual(session.calls, 2)
        self.assertEqual(len(sleeps), 1)

    def test_data_too_large_is_not_retried_identically(self):
        session = FakeSession([
            FakeResponse(
                500,
                {
                    "error": {
                        "message": (
                            "Please reduce the amount of data you're asking for, "
                            "then retry your request"
                        ),
                        "code": 1,
                    }
                },
            ),
        ])
        sleeps = []
        client = MetaAdsClient(
            api_version="v25.0",
            access_token="token",
            max_retries=5,
            sleep_func=sleeps.append,
            session=session,
        )
        with self.assertRaises(MetaApiError) as context:
            client.get_json("https://example.invalid")
        self.assertTrue(context.exception.data_too_large)
        self.assertEqual(session.calls, 1)
        self.assertEqual(sleeps, [])

    def test_rinci_insights_are_split_per_day(self):
        session = FakeSession([
            FakeResponse(200, {"data": [{"date_start": "2026-08-01"}]}),
            FakeResponse(200, {"data": [{"date_start": "2026-08-02"}]}),
            FakeResponse(200, {"data": [{"date_start": "2026-08-03"}]}),
        ])
        client = MetaAdsClient(
            api_version="v25.0",
            access_token="token",
            max_retries=0,
            session=session,
        )
        rows, _ = client.fetch_all_insights(
            account_id="123",
            mode="rinci",
            since="2026-08-01",
            until="2026-08-03",
        )
        self.assertEqual(len(rows), 3)
        ranges = [
            json.loads(call[1]["params"]["time_range"])
            for call in session.requests
        ]
        self.assertEqual(
            ranges,
            [
                {"since": "2026-08-01", "until": "2026-08-01"},
                {"since": "2026-08-02", "until": "2026-08-02"},
                {"since": "2026-08-03", "until": "2026-08-03"},
            ],
        )


class DailyResultsTests(unittest.TestCase):
    def test_daily_result_request_has_no_hourly_breakdown(self):
        session = FakeSession([
            FakeResponse(200, {"data": [{
                "ad_id": "52510676405480",
                "date_start": "2026-07-01",
                "results": [{
                    "indicator": (
                        "actions:onsite_conversion."
                        "messaging_conversation_started_7d"
                    ),
                    "values": [{"value": "9"}],
                }],
            }]}),
        ])
        client = MetaAdsClient(
            api_version="v25.0",
            access_token="token",
            max_retries=0,
            session=session,
        )
        rows, fields = client.fetch_daily_result_insights(
            account_id="123",
            since="2026-07-01",
            until="2026-07-01",
        )
        self.assertEqual(rows[0]["ad_id"], "52510676405480")
        self.assertIn("results", fields)
        params = session.requests[0][1]["params"]
        self.assertEqual(params["time_increment"], "1")
        self.assertNotIn("breakdowns", params)

    def test_daily_result_written_once_across_hourly_rows(self):
        class Client:
            def fetch_adsets(self, account_id):
                return {}

            def fetch_all_insights(self, account_id, mode, since, until):
                return ([
                    {
                        "ad_id": "52510676405480",
                        "ad_name": "Iklan",
                        "date_start": "2026-07-01",
                        "hourly_stats_aggregated_by_advertiser_time_zone": (
                            "10:00:00 - 10:59:59"
                        ),
                        "impressions": "344",
                        "inline_link_clicks": "5",
                        "inline_link_click_ctr": "1.45",
                        "spend": "4932",
                    },
                    {
                        "ad_id": "52510676405480",
                        "ad_name": "Iklan",
                        "date_start": "2026-07-01",
                        "hourly_stats_aggregated_by_advertiser_time_zone": (
                            "03:00:00 - 03:59:59"
                        ),
                        "impressions": "13",
                        "inline_link_clicks": "1",
                        "inline_link_click_ctr": "7.69",
                        "spend": "96",
                    },
                ], [])

            def fetch_daily_result_insights(self, account_id, since, until):
                return ([{
                    "ad_id": "52510676405480",
                    "date_start": "2026-07-01",
                    "spend": "9000",
                    "optimization_goal": "CONVERSATIONS",
                    "results": [{
                        "indicator": (
                            "actions:onsite_conversion."
                            "messaging_conversation_started_7d"
                        ),
                        "values": [{"value": "9"}],
                    }],
                }], [])

        rows = process_rinci_account(
            client=Client(),
            account_id="123",
            cabang="CAB",
            bisnis="GADAI",
            account_info={"name": "Akun", "currency": "IDR"},
            since="2026-07-01",
            until="2026-07-01",
            status_filter="semua",
        )
        self.assertEqual([row["Jam mulai"] for row in rows], [3, 10])
        self.assertEqual(rows[0]["Hasil"], "9")
        self.assertEqual(rows[0]["Biaya per hasil"], "1000.0")
        self.assertEqual(rows[1]["Hasil"], "")
        self.assertEqual(rows[1]["Biaya per hasil"], "")
        self.assertEqual(sum(float(row["Hasil"] or 0) for row in rows), 9.0)


class FakeMetaClient:
    created_tokens = []

    def __init__(self, *args, **kwargs):
        self.__class__.created_tokens.append(kwargs.get("access_token"))

    def get_account_information(self, account_id):
        return {
            "name": "Akun Test",
            "currency": "IDR",
            "timezone_name": "Asia/Jakarta",
        }

    def fetch_adsets(self, account_id):
        return {
            "2001": {
                "id": "2001",
                "name": "Set",
                "attribution_spec": [
                    {"event_type": "CLICK_THROUGH", "window_days": 7}
                ],
            }
        }

    def fetch_all_insights(self, account_id, mode, since, until):
        return ([{
            "campaign_id": "1001",
            "campaign_name": "Kampanye",
            "adset_id": "2001",
            "adset_name": "Set",
            "ad_id": "3001",
            "ad_name": 'Postingan: "Lagi butuh dana cepat tapi nggak mau"',
            "date_start": since,
            "hourly_stats_aggregated_by_advertiser_time_zone": (
                "18:00:00 - 18:59:59"
            ),
            "impressions": "100",
            "inline_link_clicks": "5",
            "inline_link_click_ctr": "5",
            "spend": "10000",
        }], [])

    def fetch_daily_result_insights(self, account_id, since, until):
        return ([{
            "ad_id": "3001",
            "date_start": since,
            "spend": "10000",
            "optimization_goal": "CONVERSATIONS",
            "results": [{
                "indicator": (
                    "actions:onsite_conversion."
                    "messaging_conversation_started_7d"
                ),
                "values": [{"value": "9", "attribution_windows": ["default"]}],
            }],
        }], [])


class RunnerTests(unittest.TestCase):
    def test_per_account_token_and_checkpoint(self):
        FakeMetaClient.created_tokens = []
        with tempfile.TemporaryDirectory() as temp_dir:
            checkpoint_path = os.path.join(temp_dir, "state.sqlite3")
            signature = {"schema_version": "2", "test": "runner"}
            store = CheckpointStore(checkpoint_path, signature, resume=False)
            try:
                result = run_accounts(
                    accounts=[
                        {
                            "account_id": "123",
                            "access_token": "secret-token-123",
                            "cabang": "CAB",
                            "bisnis": "GADAI",
                        },
                        {
                            "account_id": "456",
                            "access_token": "secret-token-456",
                            "cabang": "CAB2",
                            "bisnis": "MIKRO",
                        },
                    ],
                    mode="rinci",
                    status_filter="semua",
                    since="2026-08-01",
                    until="2026-08-05",
                    checkpoint=store,
                    api_version="v25.0",
                    max_retries=0,
                    retry_base_delay=0,
                    retry_max_delay=0,
                    client_factory=FakeMetaClient,
                )
                self.assertEqual(
                    FakeMetaClient.created_tokens,
                    ["secret-token-123", "secret-token-456"],
                )
                self.assertEqual(len(result.rows), 2)
                row = result.rows[0]
                self.assertEqual(row["ID ACCOUNT"], "123")
                self.assertEqual(row["CABANG"], "CAB")
                self.assertEqual(row["BISNIS"], "GADAI")
                self.assertEqual(row["ID kampanye"], "1001")
                self.assertEqual(row["ID set iklan"], "2001")
                self.assertEqual(row["ID iklan"], "3001")
                self.assertEqual(row["Jam mulai"], 18)
                self.assertEqual(row["Kategori waktu"], "Malam")
                self.assertEqual(row["Hasil"], "9")
                self.assertEqual(
                    row["Nama iklan"],
                    "Lagi butuh dana cepat tapi nggak mau",
                )
                self.assertNotIn("TOKEN", row)
                self.assertEqual(store.completed_accounts(), {"123", "456"})

                tables = store.connection.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                ).fetchall()
                self.assertNotIn("token", " ".join(str(item) for item in tables).lower())
                account_columns = [
                    item[1]
                    for item in store.connection.execute("PRAGMA table_info(accounts)")
                ]
                self.assertNotIn("token", account_columns)
            finally:
                store.close()


class CheckpointTests(unittest.TestCase):
    def test_roundtrip_and_resume(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "checkpoint.sqlite3")
            signature = {"test": "signature", "schema_version": "2"}
            store = CheckpointStore(path, signature, resume=False)
            store.save_account_result(
                account_id="123",
                cabang="CAB",
                bisnis="GADAI",
                rows=[{"ID iklan": "456", "Nama iklan": "Test"}],
                errors=[],
                status="completed",
            )
            store.close()

            resumed = CheckpointStore(path, signature, resume=True)
            self.assertEqual(resumed.completed_accounts(), {"123"})
            self.assertEqual(resumed.load_rows()[0]["ID iklan"], "456")
            self.assertEqual(resumed.account_summary()[0]["bisnis"], "GADAI")
            resumed.close()

            with self.assertRaises(CheckpointMismatchError):
                CheckpointStore(path, {"test": "different"}, resume=True)


class CliIntegrationTests(unittest.TestCase):
    def test_cli_end_to_end_without_network(self):
        FakeMetaClient.created_tokens = []
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = os.path.join(temp_dir, "accounts.xlsx")
            output_path = os.path.join(temp_dir, "report.xlsx")
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(list(INPUT_HEADERS))
            sheet.append(["123", "token-dari-excel", "CAB", "GADAI"])
            workbook.save(input_path)
            workbook.close()

            with patch("meta_ads_tool.cli.MetaAdsClient", FakeMetaClient):
                code = cli_main([
                    "--mode",
                    "rinci",
                    "--status",
                    "semua",
                    "--since",
                    "2026-08-01",
                    "--until",
                    "2026-08-05",
                    "--output",
                    output_path,
                    input_path,
                ])

            self.assertEqual(code, 0)
            self.assertEqual(FakeMetaClient.created_tokens, ["token-dari-excel"])
            self.assertTrue(os.path.isfile(output_path))
            self.assertFalse(os.path.exists(output_path + ".checkpoint.sqlite3"))
            result_book = load_workbook(output_path, read_only=True)
            try:
                sheet = result_book["Rinci Iklan"]
                self.assertEqual(sheet.max_row, 2)
                headers = [cell.value for cell in sheet[1]]
                self.assertIn("CABANG", headers)
                self.assertIn("BISNIS", headers)
                self.assertNotIn("TOKEN", headers)
            finally:
                result_book.close()


class ExportTests(unittest.TestCase):
    @staticmethod
    def sample_row():
        return {
            "ID ACCOUNT": "123",
            "Nama akun": "=nama-berbahaya",
            "ID kampanye": "1001",
            "Nama kampanye": "Kampanye",
            "ID set iklan": "2001",
            "Nama set iklan": "Set",
            "ID iklan": "3001",
            "Nama iklan": "Iklan",
            "Tanggal": "2026-08-05",
            "Waktu (zona waktu akun iklan)": "07:00 - 07:59",
            "Jam mulai": 7,
            "Kategori waktu": "Pagi",
            "Pengaturan atribusi": "7d_click",
            "Mulai": "2026-08-01T00:00:00+0700",
            "Berakhir": "2026-08-31T23:59:59+0700",
            "Impresi": "100",
            "Klik tautan": "5",
            "CTR klik tautan": 5.0,
            "Biaya per klik tautan": "1000",
            "Hasil": "2",
            "Biaya per hasil": "2500",
            "Jumlah yang dibelanjakan": "5000",
            "Mata uang": "IDR",
            "Awal pelaporan": "2026-08-01",
            "Akhir pelaporan": "2026-08-05",
            "CABANG": "CAB",
            "BISNIS": "GADAI",
        }

    def test_xlsx_and_csv(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            xlsx = os.path.join(temp_dir, "report.xlsx")
            csv_path = os.path.join(temp_dir, "report.csv")
            row = self.sample_row()

            export_xlsx(
                output_file=xlsx,
                mode="rinci",
                rows=[row],
                errors=[],
                input_file=xlsx + ".input.xlsx",
                api_version="v25.0",
                since="2026-08-01",
                until="2026-08-05",
                status_filter="semua",
                date_source="manual",
                image_width=160,
                image_height=100,
                embed_images=False,
            )
            workbook = load_workbook(xlsx, data_only=False)
            try:
                sheet = workbook["Rinci Iklan"]
                headers = [cell.value for cell in sheet[1]]
                self.assertIn("ID kampanye", headers)
                self.assertIn("Jam mulai", headers)
                self.assertIn("Kategori waktu", headers)
                self.assertIn("CABANG", headers)
                self.assertIn("BISNIS", headers)
                self.assertNotIn("TOKEN", headers)
                name_col = headers.index("Nama akun") + 1
                self.assertEqual(sheet.cell(2, name_col).value, "'=nama-berbahaya")
            finally:
                workbook.close()

            result = export_csv(
                output_file=csv_path,
                mode="rinci",
                rows=[row],
                errors=[],
                input_file="input.xlsx",
                api_version="v25.0",
                since="2026-08-01",
                until="2026-08-05",
                status_filter="semua",
                date_source="manual",
                process_date="2026-08-05",
            )
            self.assertTrue(os.path.isfile(result.report_file))
            self.assertTrue(os.path.isfile(result.error_file))
            self.assertTrue(os.path.isfile(result.info_file))
            with open(csv_path, "r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.DictReader(handle)
                self.assertEqual(reader.fieldnames, CSV_REPORT_COLUMNS)
                loaded = next(reader)
                self.assertEqual(loaded["Nama akun"], "'=nama-berbahaya")
                self.assertEqual(loaded["ID iklan"], "3001")
                self.assertEqual(
                    loaded["Waktu (zona waktu akun iklan)"],
                    "07:00 - 07:59",
                )
                self.assertEqual(loaded["CABANG"], "CAB")
                self.assertEqual(loaded["BISNIS"], "GADAI")
                self.assertEqual(loaded["Tanggal proses"], "2026-08-05")
                self.assertNotIn("TOKEN", loaded)
                self.assertNotIn("ID set iklan", loaded)
                self.assertNotIn("Jam mulai", loaded)
                self.assertNotIn("Kategori waktu", loaded)


if __name__ == "__main__":
    unittest.main()
