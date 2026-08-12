"""XLSX and CSV exporters with spreadsheet-injection protection."""

import csv
import hashlib
import io
import os
import tempfile
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Dict, List, Sequence, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
try:
    from PIL import Image as PILImage
except ImportError:  # Image embedding is optional
    PILImage = None  # type: ignore

from .constants import (
    CSV_REPORT_COLUMNS,
    ERROR_COLUMNS,
    REKAP_COLUMNS,
    RINCI_COLUMNS,
)
from .logging_utils import current_date_text, warning


@dataclass
class ExportResult:
    report_file: str
    error_file: str = ""
    info_file: str = ""


_NUMERIC_COLUMNS = {
    "Impresi",
    "Jangkauan",
    "Frekuensi",
    "Klik tautan",
    "Jam mulai",
    "Biaya per klik tautan",
    "Hasil",
    "Biaya per hasil",
    "Jumlah yang dibelanjakan",
}

_ID_TEXT_COLUMNS = {
    "ID ACCOUNT",
    "ID kampanye",
    "ID set iklan",
    "ID iklan",
    "Waktu (zona waktu akun iklan)",
}

_DATE_COLUMNS = {
    "Tanggal",
    "Tanggal proses",
    "Awal pelaporan",
    "Akhir pelaporan",
}

_DATETIME_COLUMNS = {
    "Mulai",
    "Berakhir",
}


def report_columns(mode: str) -> List[str]:
    """Kolom lengkap untuk workbook XLSX."""
    return list(REKAP_COLUMNS if mode == "rekap" else RINCI_COLUMNS)


def csv_report_columns() -> List[str]:
    """Kolom ringkas CSV untuk integrasi lanjutan."""
    return list(CSV_REPORT_COLUMNS)


def sanitize_spreadsheet_text(value: Any) -> str:
    """Prevent Excel/CSV formula execution for untrusted text fields."""
    text = "" if value is None else str(value)
    probe = text.lstrip()
    if probe.startswith(("=", "+", "-", "@", "\t", "\r", "\n")):
        return "'" + text
    return text


def to_number(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        number = float(text)
    except (TypeError, ValueError):
        return sanitize_spreadsheet_text(value)
    if number.is_integer():
        return int(number)
    return number


def to_excel_date(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S%z"):
        try:
            parsed = datetime.strptime(text, fmt)
            if "T" in text:
                return parsed.replace(tzinfo=None)
            return parsed.date()
        except (TypeError, ValueError):
            continue
    return sanitize_spreadsheet_text(value)


def excel_value(column_name: str, value: Any) -> Any:
    if column_name in _ID_TEXT_COLUMNS:
        return sanitize_spreadsheet_text(value)
    if column_name in _DATE_COLUMNS or column_name in _DATETIME_COLUMNS:
        return to_excel_date(value)
    if column_name in _NUMERIC_COLUMNS:
        return to_number(value)
    if column_name == "CTR klik tautan":
        number = to_number(value)
        if isinstance(number, (int, float)):
            return float(number) / 100.0
        return number
    return sanitize_spreadsheet_text(value)


def csv_value(column_name: str, value: Any) -> Any:
    if column_name in _NUMERIC_COLUMNS or column_name == "CTR klik tautan":
        return to_number(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return sanitize_spreadsheet_text(value)


def _style_header(worksheet: Any, columns: Sequence[str], row_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bottom_border = Border(bottom=Side(style="thin", color="A6A6A6"))

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 36
    worksheet.auto_filter.ref = "A1:{}{}".format(
        get_column_letter(len(columns)),
        max(row_count + 1, 1),
    )
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = bottom_border


def _set_report_widths(worksheet: Any, columns: Sequence[str]) -> None:
    widths_by_name = {
        "ID ACCOUNT": 22,
        "Nama akun": 28,
        "ID kampanye": 22,
        "Nama kampanye": 34,
        "Tujuan kampanye": 22,
        "ID set iklan": 22,
        "Nama set iklan": 34,
        "Optimasi": 24,
        "ID iklan": 22,
        "Nama iklan": 34,
        "Status efektif": 18,
        "Status konfigurasi": 18,
        "Gambar iklan": 25,
        "URL gambar": 42,
        "Tanggal": 13,
        "Waktu (zona waktu akun iklan)": 24,
        "Jam mulai": 12,
        "Kategori waktu": 16,
        "Pengaturan atribusi": 24,
        "Mulai": 24,
        "Berakhir": 24,
        "Impresi": 14,
        "Jangkauan": 14,
        "Frekuensi": 13,
        "Klik tautan": 14,
        "CTR klik tautan": 17,
        "Biaya per klik tautan": 20,
        "Hasil": 14,
        "Biaya per hasil": 18,
        "Jumlah yang dibelanjakan": 24,
        "Mata uang": 12,
        "Awal pelaporan": 16,
        "Akhir pelaporan": 16,
        "CABANG": 18,
        "BISNIS": 20,
    }
    for index, column_name in enumerate(columns, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = widths_by_name.get(
            column_name,
            18,
        )


def _apply_number_formats(
    worksheet: Any,
    columns: Sequence[str],
    row_count: int,
) -> None:
    column_index = {name: index + 1 for index, name in enumerate(columns)}

    for row_number in range(2, row_count + 2):
        for name in _ID_TEXT_COLUMNS:
            if name in column_index:
                worksheet.cell(row_number, column_index[name]).number_format = "@"
        for name in _DATE_COLUMNS:
            if name in column_index:
                worksheet.cell(row_number, column_index[name]).number_format = "yyyy-mm-dd"
        for name in _DATETIME_COLUMNS:
            if name in column_index:
                worksheet.cell(row_number, column_index[name]).number_format = (
                    "yyyy-mm-dd hh:mm:ss"
                )
        for name in ("Impresi", "Jangkauan", "Klik tautan", "Jam mulai"):
            if name in column_index:
                worksheet.cell(row_number, column_index[name]).number_format = "#,##0"
        for name in (
            "Frekuensi",
            "Biaya per klik tautan",
            "Hasil",
            "Biaya per hasil",
            "Jumlah yang dibelanjakan",
        ):
            if name in column_index:
                worksheet.cell(row_number, column_index[name]).number_format = "#,##0.00"
        if "CTR klik tautan" in column_index:
            worksheet.cell(
                row_number,
                column_index["CTR klik tautan"],
            ).number_format = "0.00%"
        for column_number in range(1, len(columns) + 1):
            worksheet.cell(row_number, column_number).alignment = Alignment(
                vertical="top",
                wrap_text=False,
            )


def _style_error_sheet(worksheet: Any, error_count: int) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 28
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 22
    worksheet.column_dimensions["E"].width = 22
    worksheet.column_dimensions["F"].width = 100
    if error_count:
        worksheet.auto_filter.ref = "A1:F{}".format(error_count + 1)


def _download_image_to_png(
    url: str,
    temp_directory: str,
    width: int,
    height: int,
    image_session: requests.Session,
) -> Tuple[str, int, int]:
    if not url.startswith(("http://", "https://")):
        raise ValueError("URL gambar tidak valid.")

    response = image_session.get(url, timeout=(10, 60), stream=True)
    response.raise_for_status()
    content = response.content
    if len(content) > 15 * 1024 * 1024:
        raise ValueError("Ukuran gambar melebihi 15 MB.")

    if PILImage is None:
        raise RuntimeError("Pillow belum terpasang; gunakan --tanpa-gambar atau install Pillow.")
    image = PILImage.open(io.BytesIO(content))
    image.load()
    if image.mode not in ("RGB", "RGBA"):
        image = image.convert("RGBA")
    image.thumbnail((width, height))

    digest = hashlib.sha256(url.encode("utf-8")).hexdigest()[:20]
    path = os.path.join(temp_directory, "{}.png".format(digest))
    image.save(path, format="PNG")
    return path, int(image.width), int(image.height)


def _embed_rekap_images(
    worksheet: Any,
    rows: List[Dict[str, Any]],
    columns: Sequence[str],
    errors: List[Dict[str, Any]],
    temp_directory: str,
    width: int,
    height: int,
) -> None:
    image_column = columns.index("Gambar iklan") + 1
    url_column = columns.index("URL gambar") + 1
    image_session = requests.Session()
    image_session.headers.update({"User-Agent": "MetaAdsReportImageFetcher/1.0"})
    cache: Dict[str, Tuple[str, int, int]] = {}

    for offset, row in enumerate(rows, start=2):
        url = str(row.get("URL gambar", "")).strip()
        url_cell = worksheet.cell(offset, url_column)
        if url:
            url_cell.hyperlink = url
            url_cell.style = "Hyperlink"
        if not url:
            worksheet.cell(offset, image_column).value = "Tidak tersedia"
            continue

        try:
            if url not in cache:
                cache[url] = _download_image_to_png(
                    url,
                    temp_directory,
                    width,
                    height,
                    image_session,
                )
            image_path, image_width, image_height = cache[url]
            excel_image = ExcelImage(image_path)
            excel_image.width = image_width
            excel_image.height = image_height
            anchor_cell = worksheet.cell(offset, image_column).coordinate
            worksheet.add_image(excel_image, anchor_cell)
            worksheet.row_dimensions[offset].height = max(78, image_height * 0.78)
        except Exception as exc:
            worksheet.cell(offset, image_column).value = "Gagal dimuat"
            errors.append({
                "ID ACCOUNT": str(row.get("ID ACCOUNT", "")),
                "CABANG": str(row.get("CABANG", "")),
                "BISNIS": str(row.get("BISNIS", "")),
                "ID iklan": str(row.get("ID iklan", "")),
                "Tahap": "Gambar",
                "Error": str(exc),
            })
            warning(
                "Gambar untuk iklan {} gagal dimuat: {}".format(
                    row.get("ID iklan", ""),
                    exc,
                )
            )


def _information_rows(
    mode: str,
    status_filter: str,
    api_version: str,
    since: str,
    until: str,
    input_file: str,
    date_source: str,
    output_format: str,
    images_enabled: bool,
) -> List[List[Any]]:
    return [
        ["Parameter", "Nilai"],
        ["Mode", mode],
        ["Status", status_filter],
        ["API Version", api_version],
        ["Awal pelaporan", since],
        ["Akhir pelaporan", until],
        ["Sumber tanggal", date_source],
        ["Format output", output_format],
        ["Gambar tertanam", "ya" if images_enabled else "tidak"],
        ["File sumber akun", os.path.abspath(input_file)],
        ["Sumber token", "Kolom TOKEN pada file input; token tidak ditulis ke output, log, atau checkpoint."],
        ["Waktu pembuatan", datetime.now().astimezone().isoformat()],
        [
            "Definisi status",
            "Status efektif dibaca saat ekspor, bukan status historis pada tanggal laporan.",
        ],
        [
            "Definisi CTR",
            "inline_link_click_ctr / CTR klik tautan, bukan CTR Semua",
        ],
        [
            "Kategori waktu",
            "Dini hari 00-04; Pagi 05-10; Siang 11-14; Sore 15-17; Malam 18-23.",
        ],
    ]


def export_xlsx(
    output_file: str,
    mode: str,
    rows: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
    input_file: str,
    api_version: str,
    since: str,
    until: str,
    status_filter: str,
    date_source: str,
    image_width: int,
    image_height: int,
    embed_images: bool,
) -> ExportResult:
    if embed_images and PILImage is None:
        raise RuntimeError("Pillow belum terpasang; gunakan --tanpa-gambar atau install Pillow.")
    columns = report_columns(mode)
    workbook = Workbook()
    report_sheet = workbook.active
    report_sheet.title = "Rekap Iklan" if mode == "rekap" else "Rinci Iklan"
    report_sheet.append(columns)
    for row in rows:
        report_sheet.append([
            excel_value(column_name, row.get(column_name, ""))
            for column_name in columns
        ])

    _style_header(report_sheet, columns, len(rows))
    _set_report_widths(report_sheet, columns)
    _apply_number_formats(report_sheet, columns, len(rows))

    output_directory = os.path.dirname(os.path.abspath(output_file))
    if output_directory:
        os.makedirs(output_directory, exist_ok=True)

    with tempfile.TemporaryDirectory(prefix="meta_ads_images_") as temp_directory:
        if mode == "rekap" and embed_images:
            _embed_rekap_images(
                report_sheet,
                rows,
                columns,
                errors,
                temp_directory,
                image_width,
                image_height,
            )

        error_sheet = workbook.create_sheet("Error")
        error_sheet.append(ERROR_COLUMNS)
        for item in errors:
            error_sheet.append([
                sanitize_spreadsheet_text(item.get(column, ""))
                for column in ERROR_COLUMNS
            ])
        _style_error_sheet(error_sheet, len(errors))

        info_sheet = workbook.create_sheet("Informasi")
        info_sheet.sheet_view.showGridLines = False
        info_rows = _information_rows(
            mode=mode,
            status_filter=status_filter,
            api_version=api_version,
            since=since,
            until=until,
            input_file=input_file,
            date_source=date_source,
            output_format="xlsx",
            images_enabled=mode == "rekap" and embed_images,
        )
        for row in info_rows:
            info_sheet.append([sanitize_spreadsheet_text(value) for value in row])
        info_sheet.column_dimensions["A"].width = 26
        info_sheet.column_dimensions["B"].width = 100
        for cell in info_sheet[1]:
            cell.fill = PatternFill("solid", fgColor="7030A0")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        workbook.save(output_file)
    workbook.close()

    # Structural verification: ensure the generated workbook can be reopened.
    verification = load_workbook(output_file, read_only=True, data_only=False)
    try:
        expected = {report_sheet.title, "Error", "Informasi"}
        if not expected.issubset(set(verification.sheetnames)):
            raise RuntimeError("Verifikasi XLSX gagal: sheet output tidak lengkap.")
    finally:
        verification.close()

    return ExportResult(report_file=os.path.abspath(output_file))


def _write_csv(path: str, columns: Sequence[str], rows: Sequence[Dict[str, Any]]) -> None:
    parent = os.path.dirname(os.path.abspath(path))
    if parent:
        os.makedirs(parent, exist_ok=True)
    with open(path, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(columns)
        for row in rows:
            writer.writerow([
                csv_value(column, row.get(column, ""))
                for column in columns
            ])


def export_csv(
    output_file: str,
    mode: str,
    rows: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
    input_file: str,
    api_version: str,
    since: str,
    until: str,
    status_filter: str,
    date_source: str,
    process_date: str = "",
) -> ExportResult:
    # CSV sengaja memakai skema tetap/ringkas. Output XLSX tetap memakai
    # kolom lengkap sesuai mode. Field yang tidak tersedia pada suatu mode
    # akan ditulis kosong.
    columns = csv_report_columns()
    resolved_process_date = process_date or current_date_text()
    csv_rows = []
    for source_row in rows:
        output_row = dict(source_row)
        output_row["Tanggal proses"] = resolved_process_date
        csv_rows.append(output_row)
    _write_csv(output_file, columns, csv_rows)

    stem, _ = os.path.splitext(output_file)
    error_file = stem + "_errors.csv"
    info_file = stem + "_info.csv"
    _write_csv(error_file, ERROR_COLUMNS, errors)

    info_rows = _information_rows(
        mode=mode,
        status_filter=status_filter,
        api_version=api_version,
        since=since,
        until=until,
        input_file=input_file,
        date_source=date_source,
        output_format="csv",
        images_enabled=False,
    )
    with open(info_file, "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        for row in info_rows:
            writer.writerow([sanitize_spreadsheet_text(value) for value in row])

    return ExportResult(
        report_file=os.path.abspath(output_file),
        error_file=os.path.abspath(error_file),
        info_file=os.path.abspath(info_file),
    )


def export_report(
    output_format: str,
    output_file: str,
    mode: str,
    rows: List[Dict[str, Any]],
    errors: List[Dict[str, Any]],
    input_file: str,
    api_version: str,
    since: str,
    until: str,
    status_filter: str,
    date_source: str,
    image_width: int,
    image_height: int,
    embed_images: bool,
) -> ExportResult:
    if output_format == "xlsx":
        return export_xlsx(
            output_file=output_file,
            mode=mode,
            rows=rows,
            errors=errors,
            input_file=input_file,
            api_version=api_version,
            since=since,
            until=until,
            status_filter=status_filter,
            date_source=date_source,
            image_width=image_width,
            image_height=image_height,
            embed_images=embed_images,
        )
    if output_format == "csv":
        return export_csv(
            output_file=output_file,
            mode=mode,
            rows=rows,
            errors=errors,
            input_file=input_file,
            api_version=api_version,
            since=since,
            until=until,
            status_filter=status_filter,
            date_source=date_source,
        )
    raise ValueError("Format output tidak didukung: {}".format(output_format))
