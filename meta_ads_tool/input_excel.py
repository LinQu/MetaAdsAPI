"""Read and validate account lists from XLSX.

Expected input columns, in this exact order:
A. ID ACCOUNT
B. TOKEN
C. CABANG
D. BISNIS
"""

import os
import re
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from .constants import INPUT_HEADERS
from .logging_utils import info, warning


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().upper().split())


def normalize_account_id(value: Any, cell_coordinate: str) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        raise ValueError(
            "Nilai ID akun pada {} berupa boolean, bukan ID akun.".format(
                cell_coordinate
            )
        )

    if isinstance(value, int):
        account_id = str(value)
    elif isinstance(value, float):
        if not value.is_integer():
            raise ValueError(
                "Nilai ID akun pada {} memiliki desimal: {}".format(
                    cell_coordinate,
                    value,
                )
            )
        account_id = "{:.0f}".format(value)
    else:
        account_id = str(value).strip()
        if account_id.startswith("'"):
            account_id = account_id[1:].strip()
        account_id = re.sub(r"^act_", "", account_id, flags=re.IGNORECASE)
        account_id = account_id.replace(" ", "")
        if "e+" in account_id.lower() or "e-" in account_id.lower():
            raise ValueError(
                "ID akun pada {} terbaca sebagai notasi ilmiah. "
                "Format kolom ID ACCOUNT sebagai Text dan masukkan ulang ID aslinya.".format(
                    cell_coordinate
                )
            )

    if not account_id:
        return ""
    if not account_id.isdigit():
        raise ValueError(
            "ID akun pada {} harus hanya berisi angka, diterima: '{}'".format(
                cell_coordinate,
                account_id,
            )
        )
    if len(account_id) > 15 and not isinstance(value, str):
        warning(
            "{} berisi ID {} sebagai angka Excel >15 digit. Presisinya mungkin "
            "sudah berubah. Format kolom ID ACCOUNT sebagai Text.".format(
                cell_coordinate,
                account_id,
            )
        )
    return account_id


def normalize_token(value: Any, cell_coordinate: str) -> str:
    """Return a token without ever including it in error messages or logs."""
    if value is None:
        return ""
    if isinstance(value, bool):
        raise ValueError(
            "TOKEN pada {} tidak valid. Simpan token sebagai teks.".format(
                cell_coordinate
            )
        )
    token = str(value).strip()
    if token.startswith("'"):
        token = token[1:].strip()
    if not token:
        return ""
    if any(character in token for character in ("\r", "\n", "\t")):
        raise ValueError(
            "TOKEN pada {} mengandung karakter baris/tab yang tidak valid.".format(
                cell_coordinate
            )
        )
    return token


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def validate_input_file(input_file: str) -> None:
    if not os.path.isfile(input_file):
        raise RuntimeError(
            "File input tidak ditemukan: {}".format(os.path.abspath(input_file))
        )
    if not input_file.lower().endswith(".xlsx"):
        raise RuntimeError("File input harus berekstensi .xlsx.")


def read_accounts_from_excel(
    input_file: str,
    input_sheet: str = "",
) -> List[Dict[str, str]]:
    """Read account credentials from the exact A:D input schema.

    Returned dictionaries contain ``account_id``, ``access_token``, ``cabang``,
    and ``bisnis``. The token is kept in memory only and is never logged.
    """
    validate_input_file(input_file)
    workbook = load_workbook(input_file, read_only=True, data_only=True)
    try:
        if input_sheet:
            if input_sheet not in workbook.sheetnames:
                raise RuntimeError(
                    "Sheet '{}' tidak ditemukan. Sheet tersedia: {}".format(
                        input_sheet,
                        ", ".join(workbook.sheetnames),
                    )
                )
            worksheet = workbook[input_sheet]
        else:
            worksheet = workbook.active

        expected_headers = list(INPUT_HEADERS)
        header_row_number: Optional[int] = None
        max_header_scan = min(30, worksheet.max_row or 30)

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=max_header_scan),
            start=1,
        ):
            first_four = [
                normalize_header(row[index].value) if index < len(row) else ""
                for index in range(4)
            ]
            if first_four == expected_headers:
                header_row_number = row_number
                break

        if header_row_number is None:
            raise RuntimeError(
                "Header input harus berada berurutan pada kolom A:D: {}. "
                "Header tersebut tidak ditemukan pada 30 baris pertama sheet '{}'.".format(
                    " | ".join(expected_headers),
                    worksheet.title,
                )
            )

        accounts_by_id: Dict[str, Dict[str, str]] = {}
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row_number + 1),
            start=header_row_number + 1,
        ):
            values = [row[index].value if index < len(row) else None for index in range(4)]
            if all(value is None or str(value).strip() == "" for value in values):
                continue

            account_cell = row[0]
            token_cell = row[1] if len(row) > 1 else None
            cabang_cell = row[2] if len(row) > 2 else None
            bisnis_cell = row[3] if len(row) > 3 else None

            account_id = normalize_account_id(
                account_cell.value,
                account_cell.coordinate,
            )
            if not account_id:
                warning(
                    "Baris {} dilewati karena ID ACCOUNT kosong.".format(row_number)
                )
                continue

            if token_cell is None:
                raise RuntimeError("Kolom TOKEN tidak tersedia pada baris {}.".format(row_number))
            access_token = normalize_token(token_cell.value, token_cell.coordinate)
            cabang = normalize_text(cabang_cell.value if cabang_cell is not None else None)
            bisnis = normalize_text(bisnis_cell.value if bisnis_cell is not None else None)

            missing = []
            if not access_token:
                missing.append("TOKEN")
            if not cabang:
                missing.append("CABANG")
            if not bisnis:
                missing.append("BISNIS")
            if missing:
                raise RuntimeError(
                    "Baris {} untuk ID ACCOUNT {} memiliki kolom wajib kosong: {}.".format(
                        row_number,
                        account_id,
                        ", ".join(missing),
                    )
                )

            current = {
                "account_id": account_id,
                "access_token": access_token,
                "cabang": cabang,
                "bisnis": bisnis,
            }
            if account_id in accounts_by_id:
                previous = accounts_by_id[account_id]
                differences = []
                if previous["access_token"] != access_token:
                    differences.append("TOKEN")
                if previous["cabang"] != cabang:
                    differences.append("CABANG")
                if previous["bisnis"] != bisnis:
                    differences.append("BISNIS")
                if differences:
                    warning(
                        "ID ACCOUNT {} muncul lebih dari sekali dengan {} berbeda; "
                        "baris pertama dipakai.".format(
                            account_id,
                            ", ".join(differences),
                        )
                    )
                continue

            accounts_by_id[account_id] = current

        accounts = list(accounts_by_id.values())
        if not accounts:
            raise RuntimeError(
                "Tidak ada ID ACCOUNT valid pada sheet '{}'.".format(
                    worksheet.title
                )
            )
        info(
            "Ditemukan {} akun dari sheet '{}'. Token dibaca per akun dan tidak ditampilkan di log.".format(
                len(accounts),
                worksheet.title,
            )
        )
        return accounts
    finally:
        workbook.close()
